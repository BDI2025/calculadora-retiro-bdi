"""
================================================================
    CALCULADORA DE RETIRO BDI
    Plataforma con 3 modos de proyección + módulo educativo
    Estética BDI: verde #137247 / cyan #17BEBB / lima #B5E61D
================================================================
Para correr local:
    pip install -r requirements.txt
    streamlit run app.py
"""

import os
from io import BytesIO
from datetime import datetime

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import requests

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.pdfgen import canvas as rl_canvas

# ============================================================
# GOOGLE FORM - LEAD CAPTURE BDI
# ============================================================
GOOGLE_FORM_ACTION = (
    "https://docs.google.com/forms/d/e/"
    "1FAIpQLSe3OGgBrW7agMmxCU0KIldNov8VWnTHUxFLbIsWInp0Uc1Mqw/formResponse"
)
GOOGLE_FORM_ENTRIES = {
    "nombre":   "entry.2093570355",
    "email":    "entry.1889827914",
    "telefono": "entry.393747225",
    "broker":   "entry.311873293",
    "canal":    "entry.497451233",
}
BDI_EMAIL = "hola@bdiconsultora.com"
BDI_WEB = "www.bdiconsultora.com"

# ============================================================
# PALETA Y CONSTANTES BDI
# ============================================================
BDI_VERDE = "#137247"
BDI_GRIS = "#323232"
BDI_CREMA = "#EFEDEA"
BDI_CYAN = "#17BEBB"
BDI_LIMA = "#B5E61D"
BDI_GRIS_SUAVE = "#5F5E5A"

LOGO_PATH_OPTIONS = [
    "assets/logo_bdi.png",
    "assets/logo_bdi.svg",
    "assets/logo_bdi.jpg",
    "assets/logo_bdi.jpeg",
]

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="BDI · Calculadora de Retiro",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# CSS PERSONALIZADO
# ============================================================
def inject_css() -> None:
    st.markdown(
        f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&family=Cormorant+Garamond:wght@600;700&family=Bebas+Neue&display=swap');

        html, body, [class*="css"] {{
            font-family: 'Poppins', system-ui, sans-serif;
        }}

        .bdi-header {{
            background: {BDI_GRIS};
            padding: 18px 24px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            gap: 18px;
            margin-bottom: 14px;
        }}
        .bdi-header-title {{
            font-family: 'Bebas Neue', Impact, sans-serif;
            font-size: 26px;
            color: {BDI_CREMA};
            letter-spacing: 1.5px;
            line-height: 1;
        }}
        .bdi-header-subtitle {{
            font-size: 11px;
            color: {BDI_LIMA};
            letter-spacing: 2px;
            margin-top: 4px;
        }}

        .kpi-card {{
            background: {BDI_CREMA};
            border-radius: 10px;
            padding: 14px;
            text-align: center;
        }}
        .kpi-label {{
            font-size: 12px;
            color: {BDI_GRIS_SUAVE};
            letter-spacing: 1.5px;
        }}
        .kpi-value {{
            font-family: 'Bebas Neue', Impact, sans-serif;
            font-size: 32px;
            letter-spacing: 1px;
            line-height: 1.1;
        }}
        .kpi-card-primary {{
            background: linear-gradient(135deg, {BDI_VERDE}, {BDI_CYAN});
            color: white;
        }}
        .kpi-card-primary .kpi-label {{
            color: {BDI_LIMA};
        }}
        .kpi-card-primary .kpi-value {{
            color: white;
        }}

        .section-title {{
            font-family: 'Bebas Neue', Impact, sans-serif;
            font-size: 22px;
            color: {BDI_VERDE};
            letter-spacing: 1.5px;
            margin: 18px 0 10px 0;
        }}

        .info-card {{
            background: {BDI_CREMA};
            border-left: 4px solid {BDI_VERDE};
            padding: 14px;
            border-radius: 8px;
            margin-bottom: 10px;
        }}

        .stTabs [data-baseweb="tab-list"] {{
            gap: 4px;
        }}
        .stTabs [data-baseweb="tab"] {{
            background: white;
            border-radius: 6px 6px 0 0;
            padding: 10px 18px;
            font-family: 'Bebas Neue', Impact, sans-serif;
            letter-spacing: 1.5px;
            font-size: 14px;
            color: {BDI_GRIS_SUAVE};
        }}
        .stTabs [aria-selected="true"] {{
            background: {BDI_CREMA};
            color: {BDI_VERDE} !important;
            border-bottom: 3px solid {BDI_VERDE};
        }}

        [data-testid="stSidebar"] {{
            background: {BDI_CREMA};
        }}

        .disclaimer {{
            background: {BDI_GRIS};
            color: {BDI_CREMA};
            padding: 12px;
            border-radius: 8px;
            font-size: 11px;
            line-height: 1.5;
            margin-top: 16px;
        }}
        .disclaimer strong {{
            color: {BDI_LIMA};
        }}

        .titular-grande {{
            text-align: center;
            margin: 14px 0 18px 0;
        }}
        .titular-grande .titulo {{
            font-size: 14px;
            color: {BDI_GRIS};
        }}
        .titular-grande .monto {{
            font-family: 'Bebas Neue', Impact, sans-serif;
            font-size: 56px;
            color: {BDI_VERDE};
            letter-spacing: 2px;
            line-height: 1.05;
        }}
        .titular-grande .pie {{
            font-size: 12px;
            color: {BDI_GRIS_SUAVE};
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# LOGO BDI - SVG DE RESPALDO (si no hay archivo en assets/)
# ============================================================
def render_logo_bdi_svg(height: int = 44, color_letras: str = "white") -> str:
    """Logo SVG de respaldo, replica el chevron compacto v2 con notch interno.
    Chevron pegado a la I de BDI (sin gap)."""
    return f"""
    <svg viewBox="0 0 175 80" style="height:{height}px;">
        <text x="0" y="64" font-family="'Cormorant Garamond', 'Playfair Display', 'Times New Roman', serif"
              font-size="72" font-weight="700" fill="{color_letras}" letter-spacing="2">BDI</text>
        <path d="M 130 18 L 162 40 L 130 62 L 130 53 L 149 40 L 130 27 Z" fill="{BDI_VERDE}"/>
    </svg>
    """


def get_logo_path() -> str | None:
    """Busca un archivo de logo real en assets/. Devuelve la ruta o None."""
    for p in LOGO_PATH_OPTIONS:
        if os.path.exists(p):
            return p
    return None


# ============================================================
# FUNCIONES DE CÁLCULO
# ============================================================
def calcular_proyeccion(
    capital_inicial: float,
    aporte_mensual: float,
    anios: int,
    retorno_anual: float,
    capitalizacion: str = "mensual",
) -> pd.DataFrame:
    """Cálculo determinístico mes a mes. Devuelve DataFrame con saldo y acumulados."""
    total_meses = anios * 12
    saldos = np.zeros(total_meses)
    aportes_acum = np.zeros(total_meses)
    intereses_acum = np.zeros(total_meses)

    saldo = float(capital_inicial)
    aporte_total = 0.0
    interes_total = 0.0

    if capitalizacion == "mensual":
        r_periodo = retorno_anual / 12
        for i in range(total_meses):
            saldo += aporte_mensual
            aporte_total += aporte_mensual
            interes = saldo * r_periodo
            saldo += interes
            interes_total += interes
            saldos[i] = saldo
            aportes_acum[i] = aporte_total
            intereses_acum[i] = interes_total
    else:  # anual
        for i in range(total_meses):
            saldo += aporte_mensual
            aporte_total += aporte_mensual
            if (i + 1) % 12 == 0:
                interes = saldo * retorno_anual
                saldo += interes
                interes_total += interes
            saldos[i] = saldo
            aportes_acum[i] = aporte_total
            intereses_acum[i] = interes_total

    df = pd.DataFrame({
        "mes": np.arange(1, total_meses + 1),
        "anio": ((np.arange(1, total_meses + 1) - 1) // 12) + 1,
        "saldo": saldos,
        "aportes_acum": aportes_acum,
        "intereses_acum": intereses_acum,
    })
    return df


def tabla_anual(df: pd.DataFrame, capital_inicial: float) -> pd.DataFrame:
    """Resume el DataFrame mensual a vista anual."""
    anual = df.groupby("anio").agg(
        saldo_final=("saldo", "last"),
        aportes_acum=("aportes_acum", "last"),
        intereses_acum=("intereses_acum", "last"),
    ).reset_index()

    anual["saldo_inicial"] = anual["saldo_final"].shift(1).fillna(capital_inicial)
    anual["aportes_anio"] = anual["aportes_acum"].diff().fillna(anual["aportes_acum"].iloc[0])
    anual["intereses_anio"] = anual["intereses_acum"].diff().fillna(anual["intereses_acum"].iloc[0])
    anual["pct_intereses"] = np.where(
        anual["aportes_anio"] > 0,
        anual["intereses_anio"] / anual["aportes_anio"] * 100,
        0.0,
    )

    return anual[["anio", "saldo_inicial", "aportes_anio", "intereses_anio", "saldo_final", "pct_intereses"]]


def simular_monte_carlo(
    capital_inicial: float,
    aporte_mensual: float,
    anios: int,
    retorno_promedio: float,
    sigma: float = 0.15,
    n_simulaciones: int = 1000,
    seed: int | None = None,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Simula n trayectorias con retornos anuales ~ N(retorno_promedio, sigma).
    Devuelve:
      - trayectorias: matriz (n_sim, total_meses+1) con saldos por mes (col 0 = capital inicial).
      - retornos_anuales: matriz (n_sim, anios) con los retornos sorteados (para clasificar
        después por sequence-of-returns).
    """
    rng = np.random.default_rng(seed)
    total_meses = anios * 12
    trayectorias = np.zeros((n_simulaciones, total_meses + 1))
    trayectorias[:, 0] = capital_inicial

    # Sortear retornos anuales con distribución normal en torno al promedio
    retornos_anuales = rng.normal(retorno_promedio, sigma, size=(n_simulaciones, anios))
    retornos_mensuales = retornos_anuales / 12

    saldos = np.full(n_simulaciones, capital_inicial, dtype=float)
    for mes in range(1, total_meses + 1):
        anio_idx = (mes - 1) // 12
        r = retornos_mensuales[:, anio_idx]
        saldos = (saldos + aporte_mensual) * (1.0 + r)
        trayectorias[:, mes] = saldos

    return trayectorias, retornos_anuales


def capital_para_meta(ingreso_mensual: float, tasa_retiro: float = 0.04) -> float:
    """Capital necesario para retirar X por mes según regla de Bengen."""
    return (ingreso_mensual * 12) / tasa_retiro


def sugerencias_para_meta(
    capital_proyectado: float,
    capital_necesario: float,
    capital_inicial: float,
    aporte_mensual: float,
    anios: int,
    retorno: float,
    capitalizacion: str = "mensual",
) -> dict | None:
    """Devuelve 3 sugerencias para cerrar la distancia a la meta."""
    if capital_proyectado >= capital_necesario:
        return None

    # Búsqueda binaria del aporte que alcanza la meta
    def saldo_final(cap_ini, aporte, n_anios):
        df = calcular_proyeccion(cap_ini, aporte, n_anios, retorno, capitalizacion)
        return df["saldo"].iloc[-1]

    # Aporte
    lo, hi = aporte_mensual, max(aporte_mensual * 50, 50000)
    while saldo_final(capital_inicial, hi, anios) < capital_necesario and hi < 1_000_000:
        hi *= 2
    for _ in range(40):
        mid = (lo + hi) / 2
        if saldo_final(capital_inicial, mid, anios) < capital_necesario:
            lo = mid
        else:
            hi = mid
    nuevo_aporte = round(hi)

    # Plazo
    nuevo_plazo = anios
    while nuevo_plazo < 80 and saldo_final(capital_inicial, aporte_mensual, nuevo_plazo) < capital_necesario:
        nuevo_plazo += 1

    # Capital inicial
    lo, hi = capital_inicial, max(capital_inicial * 50, 100_000)
    while saldo_final(hi, aporte_mensual, anios) < capital_necesario and hi < 100_000_000:
        hi *= 2
    for _ in range(40):
        mid = (lo + hi) / 2
        if saldo_final(mid, aporte_mensual, anios) < capital_necesario:
            lo = mid
        else:
            hi = mid
    nuevo_capital = round(hi)

    return {
        "aporte": nuevo_aporte,
        "aporte_extra": max(0, nuevo_aporte - aporte_mensual),
        "plazo": nuevo_plazo,
        "plazo_extra": max(0, nuevo_plazo - anios),
        "capital_inicial": nuevo_capital,
        "capital_extra": max(0, nuevo_capital - capital_inicial),
    }


# ============================================================
# EXPORTAR EXCEL CON MARCA BDI
# ============================================================
def exportar_excel(
    capital_inicial: float,
    aporte_mensual: float,
    anios: int,
    retorno: float,
    capitalizacion: str,
    df_anual: pd.DataFrame,
    capital_final: float,
    nombre_cliente: str = "Cliente BDI",
) -> bytes:
    """Genera un Excel con marca BDI: portada + tabla anual + datos gráfico."""
    output = BytesIO()
    wb = openpyxl.Workbook()

    verde_fill = PatternFill(start_color="137247", end_color="137247", fill_type="solid")
    crema_fill = PatternFill(start_color="EFEDEA", end_color="EFEDEA", fill_type="solid")
    gris_fill = PatternFill(start_color="323232", end_color="323232", fill_type="solid")

    # ---------- Hoja 1 : Portada ----------
    ws = wb.active
    ws.title = "Portada"

    ws.merge_cells("A1:F4")
    ws["A1"] = "BDI ▶  CALCULADORA DE RETIRO"
    ws["A1"].fill = gris_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=22, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, 5):
        ws.row_dimensions[r].height = 24

    ws["A6"] = "CLIENTE"
    ws["A6"].font = Font(color="5F5E5A", size=10, name="Calibri")
    ws["B6"] = nombre_cliente
    ws["B6"].font = Font(size=12, name="Calibri", bold=True)

    ws["D6"] = "FECHA DE EMISIÓN"
    ws["D6"].font = Font(color="5F5E5A", size=10, name="Calibri")
    ws["E6"] = datetime.now().strftime("%d/%m/%Y")
    ws["E6"].font = Font(size=12, name="Calibri")

    ws["A8"] = "SUPUESTOS DEL MODELO"
    ws["A8"].font = Font(color="137247", bold=True, size=14, name="Calibri")

    supuestos = [
        ("Capital inicial", f"USD {capital_inicial:,.0f}"),
        ("Aporte mensual", f"USD {aporte_mensual:,.0f}"),
        ("Plazo", f"{anios} años"),
        ("Retorno anual", f"{retorno*100:.2f}%"),
        ("Capitalización", capitalizacion.capitalize()),
        ("Moneda", "USD nominal"),
    ]
    for i, (label, value) in enumerate(supuestos):
        r = 9 + i
        ws.cell(row=r, column=1, value=label).font = Font(color="5F5E5A", size=11, name="Calibri")
        ws.cell(row=r, column=1).fill = crema_fill
        ws.cell(row=r, column=2, value=value).font = Font(color="137247", bold=True, size=12, name="Calibri")
        ws.cell(row=r, column=2).fill = crema_fill

    ws["A17"] = "RESULTADOS PROYECTADOS"
    ws["A17"].font = Font(color="137247", bold=True, size=14, name="Calibri")

    aportes_totales = float(df_anual["aportes_anio"].sum())
    intereses_totales = float(df_anual["intereses_anio"].sum())

    ws["A18"] = "Capital final"
    ws["A18"].font = Font(color="5F5E5A", size=11, name="Calibri")
    ws["B18"] = f"USD {capital_final:,.0f}"
    ws["B18"].font = Font(color="137247", bold=True, size=18, name="Calibri")

    ws["A19"] = "Aportes totales"
    ws["A19"].font = Font(color="5F5E5A", size=11, name="Calibri")
    ws["B19"] = f"USD {aportes_totales:,.0f}"
    ws["B19"].font = Font(color="17BEBB", bold=True, size=14, name="Calibri")

    ws["A20"] = "Intereses ganados"
    ws["A20"].font = Font(color="5F5E5A", size=11, name="Calibri")
    ws["B20"] = f"USD {intereses_totales:,.0f}"
    ws["B20"].font = Font(color="137247", bold=True, size=14, name="Calibri")

    ws.merge_cells("A23:F26")
    ws["A23"] = (
        "DISCLAIMER: Documento generado por la calculadora BDI con fines educativos. "
        "No constituye asesoramiento financiero personalizado. Las proyecciones asumen "
        "rendimiento constante, sin inflación, impuestos ni comisiones, y no garantizan "
        "resultados futuros. Consulte con su asesor BDI antes de tomar decisiones de inversión."
    )
    ws["A23"].fill = gris_fill
    ws["A23"].font = Font(color="EFEDEA", size=10, name="Calibri", italic=True)
    ws["A23"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_letter in "ABCDEF":
        ws.column_dimensions[col_letter].width = 22

    # ---------- Hoja 2 : Tabla anual ----------
    ws2 = wb.create_sheet("Tabla anual")
    headers = ["Año", "Saldo inicial", "Aportes del año", "Intereses ganados", "Saldo final", "% Int. vs Aporte"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = verde_fill
        cell.font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center")

    for i, row in df_anual.reset_index(drop=True).iterrows():
        r = i + 2
        ws2.cell(row=r, column=1, value=int(row["anio"]))
        ws2.cell(row=r, column=2, value=float(row["saldo_inicial"]))
        ws2.cell(row=r, column=3, value=float(row["aportes_anio"]))
        ws2.cell(row=r, column=4, value=float(row["intereses_anio"]))
        ws2.cell(row=r, column=5, value=float(row["saldo_final"]))
        ws2.cell(row=r, column=6, value=float(row["pct_intereses"]))
        for c in [2, 3, 4, 5]:
            ws2.cell(row=r, column=c).number_format = '"USD" #,##0'
        ws2.cell(row=r, column=6).number_format = '0.0"%"'
        if r % 2 == 0:
            for c in range(1, 7):
                ws2.cell(row=r, column=c).fill = crema_fill

    for col_idx in range(1, 7):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# ENVIAR LEAD AL GOOGLE FORM (POST silencioso)
# ============================================================
def enviar_lead_a_google_form(
    nombre: str,
    email: str,
    telefono: str,
    broker: str,
    canal: str,
) -> bool:
    """Envía los datos al Google Form de captura de leads. Devuelve True si OK."""
    payload = {
        GOOGLE_FORM_ENTRIES["nombre"]: nombre,
        GOOGLE_FORM_ENTRIES["email"]: email,
        GOOGLE_FORM_ENTRIES["telefono"]: telefono,
        GOOGLE_FORM_ENTRIES["broker"]: broker,
        GOOGLE_FORM_ENTRIES["canal"]: canal,
    }
    try:
        r = requests.post(GOOGLE_FORM_ACTION, data=payload, timeout=8)
        return r.status_code in (200, 302)
    except Exception:
        return False


# ============================================================
# REPORTE PDF CON MARCA BDI (6 PÁGINAS)
# ============================================================
def _draw_logo_bdi(c: rl_canvas.Canvas, x: float, y: float, height: float = 8 * mm,
                   color_letras: str = "white") -> None:
    """Dibuja el logo BDI (texto serif + chevron verde) en el canvas."""
    text_color = HexColor("#FFFFFF") if color_letras == "white" else HexColor("#000000")
    c.setFillColor(text_color)
    font_size = height * 1.25
    c.setFont("Times-Bold", font_size)
    c.drawString(x, y, "BDI")
    text_w = c.stringWidth("BDI", "Times-Bold", font_size)
    chx = x + text_w + height * 0.15
    h = height
    c.setFillColor(HexColor(BDI_VERDE))
    p = c.beginPath()
    p.moveTo(chx, y + h * 0.95)
    p.lineTo(chx + h * 0.95, y + h * 0.4)
    p.lineTo(chx, y - h * 0.15)
    p.lineTo(chx, y + h * 0.10)
    p.lineTo(chx + h * 0.55, y + h * 0.4)
    p.lineTo(chx, y + h * 0.70)
    p.close()
    c.drawPath(p, fill=1, stroke=0)


def _draw_page_header(c: rl_canvas.Canvas, w: float, h: float, seccion: str, num: int, total: int) -> None:
    """Header chico en cada página interior con logo y número de sección."""
    c.setFillColor(HexColor(BDI_VERDE))
    c.setLineWidth(0.5)
    c.line(20 * mm, h - 25 * mm, w - 20 * mm, h - 25 * mm)
    _draw_logo_bdi(c, 20 * mm, h - 22 * mm, height=5 * mm, color_letras="black")
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(w - 20 * mm, h - 22 * mm, f"{seccion} · {num} / {total}")


def _draw_page_footer(c: rl_canvas.Canvas, w: float, h: float) -> None:
    """Footer chico de cada página con disclaimer mínimo y datos de contacto BDI."""
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(20 * mm, 12 * mm,
                 "Documento BDI · No constituye asesoramiento financiero personalizado")
    c.drawRightString(w - 20 * mm, 12 * mm, f"{BDI_WEB} · {BDI_EMAIL}")


def _draw_kpi_box(c: rl_canvas.Canvas, x: float, y: float, ww: float, hh: float,
                  label: str, value: str, color_value: str = None,
                  bg: str = None, primary: bool = False) -> None:
    """Caja de KPI estilo BDI."""
    if primary:
        c.setFillColor(HexColor(BDI_VERDE))
        c.roundRect(x, y, ww, hh, 5, fill=1, stroke=0)
        label_color = HexColor(BDI_LIMA)
        value_color = HexColor("#FFFFFF")
    else:
        c.setFillColor(HexColor(bg or BDI_CREMA))
        c.roundRect(x, y, ww, hh, 5, fill=1, stroke=0)
        label_color = HexColor(BDI_GRIS_SUAVE)
        value_color = HexColor(color_value or BDI_VERDE)
    c.setFillColor(label_color)
    c.setFont("Helvetica", 7)
    c.drawCentredString(x + ww / 2, y + hh - 6, label)
    c.setFillColor(value_color)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(x + ww / 2, y + hh / 2 - 7, value)


def generar_pdf_report(
    nombre_cliente: str,
    email_cliente: str,
    capital_inicial: float,
    aporte_mensual: float,
    anios: int,
    retorno: float,
    capitalizacion: str,
    df_anual: pd.DataFrame,
    capital_final: float,
    ingreso_meta_mensual: float = 2500,
    tasa_retiro: float = 0.04,
    mc_pesimista: float | None = None,
    mc_base: float | None = None,
    mc_optimista: float | None = None,
    mc_prob_meta: float | None = None,
    mc_prob_base: float | None = None,
) -> bytes:
    """Genera el reporte PDF de 6 páginas con marca BDI."""
    output = BytesIO()
    c = rl_canvas.Canvas(output, pagesize=A4)
    w, h = A4
    fecha = datetime.now().strftime("%d de %B de %Y")

    aportes_totales = float(df_anual["aportes_anio"].sum())
    intereses_totales = float(df_anual["intereses_anio"].sum())
    cap_necesario = (ingreso_meta_mensual * 12) / tasa_retiro
    pct_meta = capital_final / cap_necesario * 100 if cap_necesario > 0 else 0

    # ================== PÁGINA 1 - PORTADA ==================
    c.setFillColor(HexColor(BDI_GRIS))
    c.rect(0, h * 0.55, w, h * 0.45, fill=1, stroke=0)
    _draw_logo_bdi(c, w / 2 - 22 * mm, h * 0.78, height=22 * mm, color_letras="white")
    c.setFillColor(HexColor(BDI_LIMA))
    c.setFont("Helvetica", 9)
    c.drawCentredString(w / 2, h * 0.70, "CONSULTORA PATRIMONIAL INTEGRAL")

    c.setFillColor(HexColor(BDI_CREMA))
    c.rect(0, 0, w, h * 0.55, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 38)
    c.drawCentredString(w / 2, h * 0.45, "PROYECCIÓN")
    c.drawCentredString(w / 2, h * 0.38, "DE RETIRO")
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 11)
    c.drawCentredString(w / 2, h * 0.32,
                        "Análisis personalizado de flujo de fondos invertidos")
    c.drawCentredString(w / 2, h * 0.295,
                        "con escenarios determinístico y Monte Carlo")

    box_w, box_h = 100 * mm, 32 * mm
    box_x, box_y = (w - box_w) / 2, h * 0.13
    c.setFillColor(HexColor("#FFFFFF"))
    c.roundRect(box_x, box_y, box_w, box_h, 4, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_VERDE))
    c.rect(box_x, box_y, 1.5 * mm, box_h, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 7)
    c.drawString(box_x + 6 * mm, box_y + box_h - 6 * mm, "CLIENTE")
    c.setFillColor(HexColor(BDI_GRIS))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(box_x + 6 * mm, box_y + box_h - 12 * mm, nombre_cliente)
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 7)
    c.drawString(box_x + 6 * mm, box_y + box_h - 19 * mm, "FECHA")
    c.setFillColor(HexColor(BDI_GRIS))
    c.setFont("Helvetica", 11)
    c.drawString(box_x + 6 * mm, box_y + box_h - 25 * mm, fecha)

    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(w / 2, 15 * mm, f"{BDI_WEB} · {BDI_EMAIL}")
    c.showPage()

    # ================== PÁGINA 2 - RESUMEN EJECUTIVO ==================
    _draw_page_header(c, w, h, "RESUMEN EJECUTIVO", 1, 6)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 38 * mm, "SUPUESTOS DEL MODELO")

    supuestos = [
        ("CAPITAL INICIAL", f"USD {capital_inicial:,.0f}"),
        ("APORTE MENSUAL", f"USD {aporte_mensual:,.0f}"),
        ("PLAZO", f"{anios} años"),
        ("RETORNO ANUAL", f"{retorno*100:.2f}%"),
        ("CAPITALIZACIÓN", capitalizacion.capitalize()),
        ("MONEDA", "USD nominal"),
    ]
    for i, (label, value) in enumerate(supuestos):
        col = i % 3
        row = i // 3
        bx = 20 * mm + col * 58 * mm
        by = h - 50 * mm - row * 18 * mm
        c.setFillColor(HexColor(BDI_CREMA))
        c.roundRect(bx, by - 12 * mm, 55 * mm, 14 * mm, 3, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_GRIS_SUAVE))
        c.setFont("Helvetica", 7)
        c.drawString(bx + 3 * mm, by - 4 * mm, label)
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(bx + 3 * mm, by - 10 * mm, value)

    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 100 * mm, "KPIS CLAVE")

    kpi_y = h - 130 * mm
    _draw_kpi_box(c, 20 * mm, kpi_y, 70 * mm, 22 * mm,
                  "CAPITAL FINAL", f"${capital_final:,.0f}", primary=True)
    _draw_kpi_box(c, 92 * mm, kpi_y, 50 * mm, 22 * mm,
                  "APORTES TOTALES", f"${aportes_totales:,.0f}", color_value=BDI_CYAN)
    _draw_kpi_box(c, 144 * mm, kpi_y, 46 * mm, 22 * mm,
                  "INTERESES", f"${intereses_totales:,.0f}", color_value=BDI_VERDE)

    # Mini-trayectoria
    chart_y = h - 175 * mm
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, chart_y + 32 * mm, "Trayectoria de la inversión")
    c.setStrokeColor(HexColor(BDI_VERDE))
    c.setLineWidth(2)
    n_pts = len(df_anual)
    pts = []
    max_val = float(df_anual["saldo_final"].max())
    for i, row in df_anual.reset_index(drop=True).iterrows():
        px = 20 * mm + (i / (n_pts - 1)) * 170 * mm
        py = chart_y + (float(row["saldo_final"]) / max_val) * 28 * mm
        pts.append((px, py))
    if len(pts) >= 2:
        path = c.beginPath()
        path.moveTo(*pts[0])
        for pt in pts[1:]:
            path.lineTo(*pt)
        c.drawPath(path, stroke=1, fill=0)

    # Disclaimer corto
    c.setFillColor(HexColor(BDI_GRIS))
    c.roundRect(20 * mm, 25 * mm, w - 40 * mm, 18 * mm, 3, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_CREMA))
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(24 * mm, 35 * mm,
                 "Este reporte resume las 3 metodologías de proyección: clásica determinística, basada en metas (Bengen 4%)")
    c.drawString(24 * mm, 31 * mm,
                 "y Monte Carlo con sequence-of-returns. Todos los supuestos son nominales en USD.")
    _draw_page_footer(c, w, h)
    c.showPage()

    # ================== PÁGINA 3 - CLÁSICA ==================
    _draw_page_header(c, w, h, "MODO CLÁSICA", 2, 6)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 38 * mm, "PROYECCIÓN DETERMINÍSTICA")

    # Pie chart simplificado a 3 barras
    capital_pct = capital_inicial / capital_final * 100
    aportes_pct = aportes_totales / capital_final * 100
    intereses_pct = intereses_totales / capital_final * 100
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, h - 50 * mm, "Composición al año %d" % anios)
    bar_y = h - 80 * mm
    bar_h = 22 * mm
    bar_w_total = 80 * mm
    cum = 0
    for label, pct, color in [
        ("Cap. ini.", capital_pct, BDI_VERDE),
        ("Aportes", aportes_pct, BDI_CYAN),
        ("Intereses", intereses_pct, BDI_LIMA),
    ]:
        seg_w = bar_w_total * pct / 100
        c.setFillColor(HexColor(color))
        c.roundRect(20 * mm + cum, bar_y, seg_w, bar_h, 2, fill=1, stroke=0)
        cum += seg_w
    # Leyenda pie
    leg_y = bar_y - 8 * mm
    for i, (label, pct, color) in enumerate([
        ("Capital inicial", capital_pct, BDI_VERDE),
        ("Aportes totales", aportes_pct, BDI_CYAN),
        ("Intereses ganados", intereses_pct, BDI_LIMA),
    ]):
        c.setFillColor(HexColor(color))
        c.rect(20 * mm, leg_y - i * 5 * mm, 3 * mm, 3 * mm, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 8)
        c.drawString(25 * mm, leg_y - i * 5 * mm,
                     f"{label} · {pct:.1f}%")

    # ¿Listo para el retiro?
    ingreso_4pct = capital_final * 0.04 / 12
    c.setFillColor(HexColor(BDI_CREMA))
    c.roundRect(110 * mm, h - 90 * mm, 80 * mm, 35 * mm, 3, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(115 * mm, h - 53 * mm, "¿LISTO PARA EL RETIRO?")
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 7)
    c.drawString(115 * mm, h - 60 * mm, "INGRESO MENSUAL (REGLA 4%)")
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 18)
    c.drawString(115 * mm, h - 70 * mm, f"${ingreso_4pct:,.0f}")
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 7)
    c.drawString(115 * mm, h - 78 * mm, f"durante {anios} años de aporte · {capital_inicial+aportes_totales:,.0f} aportados")

    # Tabla anual años clave
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, h - 130 * mm, "TABLA ANUAL · AÑOS CLAVE")

    headers = ["Año", "Saldo ini.", "Aportes", "Intereses", "Saldo final"]
    col_x = [20, 50, 85, 120, 155]
    table_y = h - 140 * mm
    c.setFillColor(HexColor(BDI_VERDE))
    c.rect(20 * mm, table_y - 5 * mm, w - 40 * mm, 7 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 9)
    for i, hdr in enumerate(headers):
        c.drawString(col_x[i] * mm, table_y - 3 * mm, hdr)

    anios_clave = sorted(set([1, max(1, anios // 3), max(1, 2 * anios // 3), anios]))
    for j, ay in enumerate(anios_clave):
        row = df_anual[df_anual["anio"] == ay]
        if row.empty:
            continue
        rr = row.iloc[0]
        ry = table_y - 12 * mm - j * 7 * mm
        if j % 2 == 0:
            c.setFillColor(HexColor(BDI_CREMA))
            c.rect(20 * mm, ry - 1 * mm, w - 40 * mm, 6 * mm, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(col_x[0] * mm, ry, str(int(rr["anio"])))
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 9)
        c.drawString(col_x[1] * mm, ry, f"${rr['saldo_inicial']:,.0f}")
        c.setFillColor(HexColor(BDI_CYAN))
        c.drawString(col_x[2] * mm, ry, f"${rr['aportes_anio']:,.0f}")
        c.setFillColor(HexColor(BDI_GRIS))
        c.drawString(col_x[3] * mm, ry, f"${rr['intereses_anio']:,.0f}")
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(col_x[4] * mm, ry, f"${rr['saldo_final']:,.0f}")

    _draw_page_footer(c, w, h)
    c.showPage()

    # ================== PÁGINA 4 - METAS ==================
    _draw_page_header(c, w, h, "MODO METAS", 3, 6)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 38 * mm, "META DE INGRESO PASIVO")

    # Hero meta
    c.setFillColor(HexColor(BDI_VERDE))
    c.roundRect(20 * mm, h - 80 * mm, w - 40 * mm, 28 * mm, 4, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_LIMA))
    c.setFont("Helvetica", 8)
    c.drawCentredString(w / 2, h - 58 * mm, "META · INGRESO PASIVO MENSUAL")
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 30)
    c.drawCentredString(w / 2, h - 70 * mm, f"${ingreso_meta_mensual:,.0f}")
    c.setFillColor(HexColor(BDI_LIMA))
    c.setFont("Helvetica", 8)
    c.drawCentredString(w / 2, h - 76 * mm, "USD por mes durante el retiro")

    # Capital necesario vs proyectado
    _draw_kpi_box(c, 20 * mm, h - 110 * mm, 80 * mm, 22 * mm,
                  "CAPITAL NECESARIO", f"${cap_necesario:,.0f}",
                  bg="#FFFFFF", color_value=BDI_VERDE)
    _draw_kpi_box(c, 110 * mm, h - 110 * mm, 80 * mm, 22 * mm,
                  "CAPITAL PROYECTADO", f"${capital_final:,.0f}",
                  color_value=BDI_GRIS_SUAVE)

    # Gauge texto
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 36)
    c.drawCentredString(w / 2, h - 135 * mm, f"{min(pct_meta, 100):.0f}%")
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica", 10)
    c.drawCentredString(w / 2, h - 142 * mm, "de tu meta cubierta")

    if pct_meta < 100:
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 12)
        c.drawString(20 * mm, h - 160 * mm, "CÓMO REDUCIR LA DISTANCIA")
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 9)
        c.drawString(20 * mm, h - 168 * mm,
                     "→ Subí el aporte mensual · Alargá el plazo · Aumentá el capital inicial")
        c.drawString(20 * mm, h - 174 * mm,
                     "Tu asesor BDI puede modelar las 3 alternativas en tu próxima reunión.")
    else:
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 12)
        c.drawString(20 * mm, h - 160 * mm, "🎯 ¡META SUPERADA!")
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 9)
        c.drawString(20 * mm, h - 168 * mm,
                     "Tu plan actual ya supera la meta de retiro. Excelente proyección.")

    _draw_page_footer(c, w, h)
    c.showPage()

    # ================== PÁGINA 5 - MONTE CARLO ==================
    _draw_page_header(c, w, h, "MONTE CARLO", 4, 6)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 38 * mm, "SEQUENCE OF RETURNS")
    c.setFillColor(HexColor(BDI_GRIS_SUAVE))
    c.setFont("Helvetica-Oblique", 9)
    c.drawString(20 * mm, h - 44 * mm,
                 "1.000 trayectorias con retorno promedio del 10% y σ=15%")

    if mc_pesimista is not None:
        _draw_kpi_box(c, 20 * mm, h - 75 * mm, 55 * mm, 22 * mm,
                      "PESIMISTA · primeros años bajos",
                      f"${mc_pesimista:,.0f}", color_value=BDI_GRIS_SUAVE)
        _draw_kpi_box(c, 78 * mm, h - 75 * mm, 60 * mm, 22 * mm,
                      "BASE · 10% FIJO",
                      f"${mc_base:,.0f}", primary=True)
        _draw_kpi_box(c, 141 * mm, h - 75 * mm, 49 * mm, 22 * mm,
                      "OPTIMISTA · primeros años altos",
                      f"${mc_optimista:,.0f}", color_value=BDI_CYAN)

        if mc_prob_meta is not None:
            c.setFillColor(HexColor(BDI_LIMA))
            c.roundRect(20 * mm, h - 100 * mm, 80 * mm, 14 * mm, 3, fill=1, stroke=0)
            c.setFillColor(HexColor(BDI_VERDE))
            c.setFont("Helvetica-Bold", 16)
            c.drawString(24 * mm, h - 95 * mm, f"{mc_prob_meta:.0f}%")
            c.setFont("Helvetica", 8)
            c.drawString(40 * mm, h - 95 * mm, "superan tu meta")

        if mc_prob_base is not None:
            c.setFillColor(HexColor(BDI_CREMA))
            c.roundRect(110 * mm, h - 100 * mm, 80 * mm, 14 * mm, 3, fill=1, stroke=0)
            c.setFillColor(HexColor(BDI_VERDE))
            c.setFont("Helvetica-Bold", 16)
            c.drawString(114 * mm, h - 95 * mm, f"{mc_prob_base:.0f}%")
            c.setFont("Helvetica", 8)
            c.drawString(130 * mm, h - 95 * mm, "superan la base 10% fija")
    else:
        c.setFillColor(HexColor(BDI_GRIS_SUAVE))
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(20 * mm, h - 70 * mm,
                     "(Para incluir Monte Carlo en el reporte, ejecutá la simulación en la Tab 3 antes de descargar.)")

    # Conclusión educativa
    c.setFillColor(HexColor(BDI_CREMA))
    c.roundRect(20 * mm, h - 145 * mm, w - 40 * mm, 28 * mm, 3, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_VERDE))
    c.rect(20 * mm, h - 145 * mm, 1.5 * mm, 28 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(25 * mm, h - 124 * mm, "CONCLUSIÓN")
    c.setFillColor(HexColor(BDI_GRIS))
    c.setFont("Helvetica", 9)
    c.drawString(25 * mm, h - 130 * mm,
                 "El orden de los retornos importa: con el mismo promedio del 10%, las trayectorias con")
    c.drawString(25 * mm, h - 134 * mm,
                 "malos años al inicio terminan distintas que las que tuvieron buenos años de entrada.")
    c.drawString(25 * mm, h - 138 * mm,
                 "Aportes mensuales sostenidos amortiguan la volatilidad (efecto dollar cost averaging).")

    _draw_page_footer(c, w, h)
    c.showPage()

    # ================== PÁGINA 6 - PRÓXIMOS PASOS + GLOSARIO + DISCLAIMER ==================
    _draw_page_header(c, w, h, "PRÓXIMOS PASOS Y GLOSARIO", 5, 6)

    # Próximos pasos
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, h - 38 * mm, "PRÓXIMOS PASOS")

    pasos = [
        ("📅 Agendar reunión con tu asesor BDI",
         "Revisar este reporte en detalle y definir vehículos de inversión concretos."),
        ("📊 Completar perfil de riesgo",
         "Tu asesor te enviará un cuestionario de tolerancia al riesgo y horizonte de inversión."),
        ("💼 Activar / revisar tu cuenta de inversión",
         "Si todavía no tenés broker, BDI te asesora en la apertura. Si ya tenés, optimizamos la cartera."),
    ]
    for i, (titulo, desc) in enumerate(pasos):
        py = h - 50 * mm - i * 15 * mm
        c.setFillColor(HexColor(BDI_CREMA))
        c.roundRect(20 * mm, py - 11 * mm, w - 40 * mm, 12 * mm, 3, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_VERDE))
        c.rect(20 * mm, py - 11 * mm, 1.5 * mm, 12 * mm, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(25 * mm, py - 4 * mm, titulo)
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 8)
        c.drawString(25 * mm, py - 9 * mm, desc)

    # Glosario
    c.setFillColor(HexColor(BDI_VERDE))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(20 * mm, h - 105 * mm, "CONCEPTOS CLAVE")

    conceptos = [
        ("INTERÉS COMPUESTO", "El rendimiento se reinvierte. FV = P(1+r/n)^(n·t)"),
        ("CAPITALIZACIÓN", "Frecuencia de reinversión. Mensual=12 veces al año."),
        ("RETORNO ESPERADO", "Promedio histórico de carteras balanceadas: 8-10%."),
        ("REGLA DEL 4%", "Bengen 1994. Capital = ingreso anual / 0.04"),
        ("MONTE CARLO", "Simula miles de futuros con retornos aleatorios."),
        ("SEQUENCE OF RETURNS", "El orden de los buenos/malos años importa."),
    ]
    for i, (titulo, desc) in enumerate(conceptos):
        col = i % 2
        row = i // 2
        bx = 20 * mm + col * 90 * mm
        by = h - 116 * mm - row * 14 * mm
        c.setFillColor(HexColor(BDI_CREMA))
        c.roundRect(bx, by - 10 * mm, 86 * mm, 11 * mm, 3, fill=1, stroke=0)
        c.setFillColor(HexColor(BDI_VERDE))
        c.setFont("Helvetica-Bold", 8)
        c.drawString(bx + 3 * mm, by - 4 * mm, titulo)
        c.setFillColor(HexColor(BDI_GRIS))
        c.setFont("Helvetica", 7)
        c.drawString(bx + 3 * mm, by - 8 * mm, desc)

    # Disclaimer + contacto
    c.setFillColor(HexColor(BDI_GRIS))
    c.roundRect(20 * mm, 20 * mm, w - 40 * mm, 35 * mm, 3, fill=1, stroke=0)
    c.setFillColor(HexColor(BDI_LIMA))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(24 * mm, 47 * mm, "⚠ DISCLAIMER BDI")
    c.setFillColor(HexColor(BDI_CREMA))
    c.setFont("Helvetica", 7)
    c.drawString(24 * mm, 42 * mm,
                 "Este reporte fue generado por la calculadora BDI con fines educativos. No constituye asesoramiento financiero")
    c.drawString(24 * mm, 39 * mm,
                 "personalizado. Las proyecciones asumen rendimiento constante (excepto en Monte Carlo) y no garantizan resultados")
    c.drawString(24 * mm, 36 * mm,
                 "futuros. Consulte con su asesor BDI antes de tomar decisiones de inversión. Cumplimiento Ley 25.326 (Argentina).")
    c.setFillColor(HexColor(BDI_LIMA))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(24 * mm, 28 * mm, BDI_EMAIL)
    c.setFillColor(HexColor(BDI_CREMA))
    c.drawString(60 * mm, 28 * mm, "·")
    c.drawString(64 * mm, 28 * mm, BDI_WEB)
    c.drawRightString(w - 24 * mm, 28 * mm, f"Documento generado: {datetime.now().strftime('%d/%m/%Y')}")

    c.showPage()
    c.save()
    output.seek(0)
    return output.getvalue()


# ============================================================
# RENDERIZADO DEL HEADER
# ============================================================
def render_header() -> None:
    logo_path = get_logo_path()
    col1, col2 = st.columns([1, 5], gap="small")
    with col1:
        if logo_path:
            st.image(logo_path, width=160)
        else:
            st.markdown(render_logo_bdi_svg(height=58, color_letras="black"), unsafe_allow_html=True)
    with col2:
        st.markdown(
            f"""
            <div style="background:{BDI_GRIS}; padding:14px 20px; border-radius:12px; height:100%;">
                <div class="bdi-header-title">CALCULADORA DE RETIRO</div>
                <div class="bdi-header-subtitle">PROYECTÁ TU FLUJO DE FONDOS INVERTIDOS</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


# ============================================================
# SIDEBAR DE INPUTS COMPARTIDOS
# ============================================================
def render_sidebar() -> tuple[float, float, int, float, str]:
    with st.sidebar:
        logo_path = get_logo_path()
        if logo_path:
            st.image(logo_path, width=140)
        else:
            st.markdown(
                f'<div style="text-align:center; margin-bottom:14px;">{render_logo_bdi_svg(height=42, color_letras=BDI_GRIS)}</div>',
                unsafe_allow_html=True,
            )

        st.markdown(f'<div class="section-title">⚙ DATOS BASE</div>', unsafe_allow_html=True)

        capital_inicial = st.number_input(
            "Capital inicial (USD)", min_value=0, value=5000, step=500, format="%d"
        )
        aporte_mensual = st.number_input(
            "Aporte mensual (USD)", min_value=0, value=200, step=50, format="%d"
        )
        anios = st.slider("Plazo (años)", min_value=1, max_value=50, value=30)
        retorno = (
            st.slider("Retorno anual esperado (%)", min_value=0.0, max_value=20.0, value=8.0, step=0.5) / 100
        )
        capitalizacion = st.selectbox("Capitalización", ["mensual", "anual"], index=0)

        st.markdown("---")
        st.markdown(
            f'<div style="font-size:10px; color:{BDI_GRIS_SUAVE}; line-height:1.5;">'
            f'Estos datos se comparten en las 4 tabs. Cada tab puede tener inputs adicionales propios.'
            f'</div>',
            unsafe_allow_html=True,
        )

        return capital_inicial, aporte_mensual, anios, retorno, capitalizacion


# ============================================================
# LEAD CAPTURE + DESCARGA PDF
# ============================================================
def render_lead_capture_y_descarga(
    capital_inicial, aporte_mensual, anios, retorno, capitalizacion,
    df_anual, capital_final,
) -> None:
    """Form de captura de leads + botón de descarga PDF tras envío."""
    # Hero
    st.markdown(
        f"""
        <div style="background:linear-gradient(135deg, {BDI_VERDE}, {BDI_CYAN}); border-radius:10px; padding:14px 18px; color:white; margin-bottom:14px;">
            <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:20px; letter-spacing:1.5px; line-height:1;">📥 OBTENÉ TU REPORTE PERSONALIZADO</div>
            <div style="font-size:12px; color:{BDI_CREMA}; margin-top:4px;">PDF de 6 páginas con tu proyección, supuestos, escenarios Monte Carlo y conceptos clave. Completá tus datos para descargarlo.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("lead_capture_form"):
        c1, c2 = st.columns(2)
        with c1:
            nombre = st.text_input("Nombre y apellido *", key="lead_nombre")
            email = st.text_input("Email *", key="lead_email")
            telefono = st.text_input("Teléfono (opcional, con código de área)", key="lead_tel")
        with c2:
            broker = st.selectbox(
                "¿Operás actualmente con algún broker? *",
                ["", "Balanz", "IOL", "BMB", "Inviu", "Otro broker", "Aún no opero"],
                key="lead_broker",
            )
            canal = st.selectbox(
                "¿Cómo nos conociste? *",
                ["", "LinkedIn", "Instagram", "YouTube",
                 "Recomendación de un amigo o conocido", "Búsqueda en Google", "Otro"],
                key="lead_canal",
            )

        st.markdown(
            f'<div style="font-size:10px; color:{BDI_GRIS_SUAVE}; margin:6px 0 10px 0; font-style:italic;">'
            f'🔒 Tus datos quedan en una base privada de BDI. No los compartimos con terceros. '
            f'Para baja escribí a <strong>{BDI_EMAIL}</strong>.</div>',
            unsafe_allow_html=True,
        )

        submitted = st.form_submit_button(
            "✓ Enviar y descargar reporte PDF", use_container_width=True,
        )

    if submitted:
        # Validaciones simples
        if not nombre or len(nombre.strip()) < 3:
            st.error("Ingresá tu nombre completo (mínimo 3 caracteres).")
            return
        if not email or "@" not in email or "." not in email.split("@")[-1]:
            st.error("Ingresá un email válido.")
            return
        if not broker:
            st.error("Indicá si operás con algún broker.")
            return
        if not canal:
            st.error("Indicá cómo nos conociste.")
            return

        # Enviar al Google Form
        ok = enviar_lead_a_google_form(
            nombre.strip(), email.strip(), (telefono or "").strip(),
            broker, canal,
        )
        if not ok:
            st.warning("No pudimos registrar tus datos en este momento, pero igual te dejamos descargar el reporte.")

        # Generar PDF (con datos Monte Carlo si están en session_state)
        mc_pes = st.session_state.get("mc_pesimista")
        mc_base = st.session_state.get("mc_base")
        mc_opt = st.session_state.get("mc_optimista")
        mc_pm = st.session_state.get("mc_prob_meta")
        mc_pb = st.session_state.get("mc_prob_base")
        meta_mensual = st.session_state.get("meta_ingreso_mensual", 2500)
        tasa_ret = st.session_state.get("tasa_retiro", 0.04)

        with st.spinner("Generando PDF..."):
            pdf_bytes = generar_pdf_report(
                nombre_cliente=nombre.strip(),
                email_cliente=email.strip(),
                capital_inicial=capital_inicial,
                aporte_mensual=aporte_mensual,
                anios=anios,
                retorno=retorno,
                capitalizacion=capitalizacion,
                df_anual=df_anual,
                capital_final=capital_final,
                ingreso_meta_mensual=meta_mensual,
                tasa_retiro=tasa_ret,
                mc_pesimista=mc_pes,
                mc_base=mc_base,
                mc_optimista=mc_opt,
                mc_prob_meta=mc_pm,
                mc_prob_base=mc_pb,
            )

        st.success("¡Listo! Tu reporte se generó. Click abajo para descargar.")
        st.download_button(
            label="⬇ DESCARGAR REPORTE PDF",
            data=pdf_bytes,
            file_name=f"Reporte_Retiro_BDI_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )


# ============================================================
# TAB 1 - CLÁSICA
# ============================================================
def tab_clasica(capital_inicial, aporte_mensual, anios, retorno, capitalizacion) -> None:
    df = calcular_proyeccion(capital_inicial, aporte_mensual, anios, retorno, capitalizacion)
    df_anual = tabla_anual(df, capital_inicial)

    capital_final = float(df["saldo"].iloc[-1])
    aportes_totales = float(df["aportes_acum"].iloc[-1])
    intereses_totales = float(df["intereses_acum"].iloc[-1])

    # Titular grande
    st.markdown(
        f"""
        <div class="titular-grande">
            <div class="titulo">Tu inversión va a valer:</div>
            <div class="monto">${capital_final:,.0f}</div>
            <div class="pie">en {anios} años</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # 3 KPIs
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">CAPITAL INICIAL</div>'
            f'<div class="kpi-value" style="color:{BDI_VERDE};">${capital_inicial:,.0f}</div></div>',
            unsafe_allow_html=True,
        )
    with col2:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">APORTES TOTALES</div>'
            f'<div class="kpi-value" style="color:{BDI_CYAN};">${aportes_totales:,.0f}</div></div>',
            unsafe_allow_html=True,
        )
    with col3:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">INTERESES GANADOS</div>'
            f'<div class="kpi-value" style="color:#739F1A;">${intereses_totales:,.0f}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br/>", unsafe_allow_html=True)

    # 2 gráficos lado a lado
    col1, col2 = st.columns([1, 1.4])

    with col1:
        st.markdown(f'<div class="section-title">COMPOSICIÓN AL AÑO {anios}</div>', unsafe_allow_html=True)
        fig_pie = go.Figure(data=[go.Pie(
            labels=["Capital inicial", "Aportes totales", "Intereses ganados"],
            values=[capital_inicial, aportes_totales, intereses_totales],
            marker=dict(colors=[BDI_VERDE, BDI_CYAN, BDI_LIMA], line=dict(color="white", width=2)),
            hole=0,
            textinfo="percent",
            textfont=dict(size=14, color="white", family="Poppins"),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>",
        )])
        fig_pie.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(t=10, b=10, l=10, r=10),
            font=dict(family="Poppins", color=BDI_GRIS),
            showlegend=True,
            legend=dict(orientation="h", y=-0.05),
            height=340,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col2:
        st.markdown(f'<div class="section-title">CRECIMIENTO AÑO POR AÑO</div>', unsafe_allow_html=True)
        capital_inicial_serie = [capital_inicial] * len(df_anual)
        aportes_acum_serie = df_anual["aportes_anio"].cumsum().tolist()
        intereses_acum_serie = df_anual["intereses_anio"].cumsum().tolist()

        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            x=df_anual["anio"], y=capital_inicial_serie,
            name="Capital inicial", marker_color=BDI_VERDE,
            hovertemplate="Año %{x}<br>Capital inicial: $%{y:,.0f}<extra></extra>",
        ))
        fig_bar.add_trace(go.Bar(
            x=df_anual["anio"], y=aportes_acum_serie,
            name="Aportes acumulados", marker_color=BDI_CYAN,
            hovertemplate="Año %{x}<br>Aportes: $%{y:,.0f}<extra></extra>",
        ))
        fig_bar.add_trace(go.Bar(
            x=df_anual["anio"], y=intereses_acum_serie,
            name="Intereses ganados", marker_color=BDI_LIMA,
            hovertemplate="Año %{x}<br>Intereses: $%{y:,.0f}<extra></extra>",
        ))
        fig_bar.update_layout(
            barmode="stack",
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(t=10, b=10, l=10, r=10),
            font=dict(family="Poppins", color=BDI_GRIS),
            xaxis=dict(title="Año", showgrid=False),
            yaxis=dict(title="USD", tickformat="$,.0f", gridcolor="#EFEDEA"),
            legend=dict(orientation="h", y=-0.2),
            height=340,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    # ¿Listo para el retiro?
    ingreso_4pct = capital_final * 0.04 / 12
    aportes_e_inicial = capital_inicial + aportes_totales
    retorno_pct = (capital_final - aportes_e_inicial) / aportes_e_inicial * 100 if aportes_e_inicial > 0 else 0

    st.markdown(f'<div class="section-title">¿LISTO PARA EL RETIRO?</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">INGRESO MENSUAL (4%)</div>'
            f'<div class="kpi-value" style="color:{BDI_VERDE};">${ingreso_4pct:,.0f}</div></div>',
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">AÑOS APORTADOS</div>'
            f'<div class="kpi-value" style="color:{BDI_VERDE};">{anios}</div></div>',
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">RETORNO ACUMULADO</div>'
            f'<div class="kpi-value" style="color:{BDI_VERDE};">+{retorno_pct:,.0f}%</div></div>',
            unsafe_allow_html=True,
        )
    with c4:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-label">CAP. AL FINAL</div>'
            f'<div class="kpi-value" style="color:{BDI_VERDE};">${capital_final:,.0f}</div></div>',
            unsafe_allow_html=True,
        )

    # Tabla anual
    with st.expander("📋 Ver tabla anual completa"):
        df_show = df_anual.copy()
        df_show.columns = ["Año", "Saldo inicial", "Aportes del año", "Intereses ganados", "Saldo final", "% Int."]
        for c in ["Saldo inicial", "Aportes del año", "Intereses ganados", "Saldo final"]:
            df_show[c] = df_show[c].apply(lambda x: f"${x:,.0f}")
        df_show["% Int."] = df_show["% Int."].apply(lambda x: f"{x:.1f}%")
        st.dataframe(df_show, hide_index=True, use_container_width=True)

    # Export · Reporte PDF con lead capture
    st.markdown("<br/>", unsafe_allow_html=True)
    render_lead_capture_y_descarga(
        capital_inicial, aporte_mensual, anios, retorno, capitalizacion,
        df_anual, capital_final,
    )


# ============================================================
# TAB 2 - METAS / GOAL-BASED
# ============================================================
def tab_metas(capital_inicial, aporte_mensual, anios, retorno, capitalizacion) -> None:
    st.markdown(f'<div class="section-title">¿CUÁNTO QUERÉS RETIRAR POR MES?</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1.1])

    with col1:
        ingreso_meta = st.number_input(
            "Meta de ingreso pasivo mensual (USD)",
            min_value=100, value=2500, step=100, format="%d",
        )
        st.markdown(
            f"""
            <div style="background:linear-gradient(135deg, {BDI_VERDE}, {BDI_CYAN}); border-radius:10px; padding:18px; text-align:center; margin:14px 0; color:white;">
                <div style="font-size:11px; color:{BDI_CREMA}; letter-spacing:1px;">META · INGRESO PASIVO MENSUAL</div>
                <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:48px; letter-spacing:2px; line-height:1;">${ingreso_meta:,.0f}</div>
                <div style="font-size:11px; color:{BDI_LIMA};">USD por mes durante el retiro</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        tasa_retiro = st.slider(
            "Tasa de retiro segura (regla de Bengen, %)",
            2.0, 6.0, 4.0, 0.5, format="%.1f",
        ) / 100

        # Persistir para el PDF
        st.session_state["meta_ingreso_mensual"] = ingreso_meta
        st.session_state["tasa_retiro"] = tasa_retiro

        cap_necesario = capital_para_meta(ingreso_meta, tasa_retiro)
        st.markdown(
            f"""
            <div class="info-card">
                <div style="font-size:11px; color:{BDI_VERDE}; font-weight:500;">📐 CAPITAL NECESARIO</div>
                <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:30px; color:{BDI_VERDE};">
                    ${cap_necesario:,.0f}
                </div>
                <div style="font-size:10px; color:{BDI_GRIS_SUAVE};">Fórmula: ingreso anual ÷ {tasa_retiro*100:.1f}%</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col2:
        df = calcular_proyeccion(capital_inicial, aporte_mensual, anios, retorno, capitalizacion)
        capital_proyectado = float(df["saldo"].iloc[-1])
        pct_meta = capital_proyectado / cap_necesario * 100 if cap_necesario > 0 else 0

        if pct_meta >= 100:
            gauge_color = BDI_VERDE
        elif pct_meta >= 70:
            gauge_color = BDI_CYAN
        else:
            gauge_color = BDI_GRIS_SUAVE

        st.markdown(f'<div class="section-title">¿LLEGÁS A TU META?</div>', unsafe_allow_html=True)

        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number",
            value=min(pct_meta, 100),
            number=dict(suffix="%", font=dict(size=44, color=BDI_VERDE, family="Poppins")),
            gauge=dict(
                axis=dict(range=[0, 100], tickwidth=1, tickcolor=BDI_GRIS),
                bar=dict(color=gauge_color, thickness=0.7),
                bgcolor=BDI_CREMA,
                borderwidth=2,
                bordercolor="white",
                steps=[
                    {"range": [0, 50], "color": "#F5F2ED"},
                    {"range": [50, 80], "color": "#EDE9DF"},
                    {"range": [80, 100], "color": "#E5DFCF"},
                ],
                threshold=dict(line=dict(color=BDI_LIMA, width=4), thickness=0.85, value=100),
            ),
        ))
        fig_gauge.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Poppins"),
            margin=dict(t=10, b=10, l=10, r=10),
            height=260,
        )
        st.plotly_chart(fig_gauge, use_container_width=True)

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(
                f'<div class="kpi-card" style="background:white; border:1px solid {BDI_VERDE};">'
                f'<div class="kpi-label">CAPITAL NECESARIO</div>'
                f'<div class="kpi-value" style="color:{BDI_VERDE};">${cap_necesario:,.0f}</div></div>',
                unsafe_allow_html=True,
            )
        with col_b:
            st.markdown(
                f'<div class="kpi-card">'
                f'<div class="kpi-label">CAPITAL PROYECTADO</div>'
                f'<div class="kpi-value" style="color:{BDI_GRIS_SUAVE};">${capital_proyectado:,.0f}</div></div>',
                unsafe_allow_html=True,
            )

    if pct_meta < 100:
        st.markdown(
            f"""
            <div style="background:linear-gradient(90deg, {BDI_VERDE}, {BDI_CYAN}); border-radius:10px; padding:14px; color:white; margin-top:14px;">
                <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:18px; letter-spacing:1.5px;">⚡ CÓMO REDUCIR LA DISTANCIA A LA META</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.spinner("Calculando sugerencias..."):
            sug = sugerencias_para_meta(
                capital_proyectado, cap_necesario, capital_inicial, aporte_mensual, anios, retorno, capitalizacion
            )
        if sug:
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(
                    f"""
                    <div class="info-card">
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE}; letter-spacing:1px;">OPCIÓN 1 · APORTE</div>
                        <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:28px; color:{BDI_VERDE};">${sug['aporte']:,.0f}/mes</div>
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE};">+${sug['aporte_extra']:,.0f} adicionales por mes</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with c2:
                st.markdown(
                    f"""
                    <div class="info-card" style="border-left-color:{BDI_CYAN};">
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE}; letter-spacing:1px;">OPCIÓN 2 · PLAZO</div>
                        <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:28px; color:{BDI_VERDE};">{sug['plazo']} años</div>
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE};">+{sug['plazo_extra']} años más</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with c3:
                st.markdown(
                    f"""
                    <div class="info-card" style="border-left-color:{BDI_LIMA};">
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE}; letter-spacing:1px;">OPCIÓN 3 · CAPITAL INICIAL</div>
                        <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:28px; color:{BDI_VERDE};">${sug['capital_inicial']:,.0f}</div>
                        <div style="font-size:12px; color:{BDI_GRIS_SUAVE};">+${sug['capital_extra']:,.0f} adicionales</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
    else:
        st.success("🎯 ¡Felicitaciones! Tu plan actual ya supera la meta de retiro.")


# ============================================================
# TAB 3 - MONTE CARLO con SEQUENCE OF RETURNS RISK
# ============================================================
def tab_montecarlo(capital_inicial, aporte_mensual, anios, retorno, capitalizacion) -> None:
    # Hero educativo
    st.markdown(
        f"""
        <div style="background:linear-gradient(90deg, {BDI_VERDE}, {BDI_CYAN}); border-radius:10px; padding:14px 18px; color:white; margin-bottom:14px;">
            <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:18px; letter-spacing:1.5px; line-height:1;">⏱ EL ORDEN DE LOS RETORNOS IMPORTA</div>
            <div style="font-size:12px; color:{BDI_CREMA}; margin-top:4px;">Un mismo promedio anual (ej. 10%) puede dar resultados muy distintos según en qué orden caigan los buenos y malos años. Comparamos 3 escenarios.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns([1, 2])

    with col1:
        st.markdown(
            f'<div style="font-size:13px; font-weight:500; color:{BDI_VERDE}; letter-spacing:1px; margin-bottom:6px;">🎲 PARÁMETROS</div>',
            unsafe_allow_html=True,
        )

        retorno_promedio = (
            st.slider("Retorno promedio anual (%)", min_value=2.0, max_value=20.0, value=10.0, step=0.5) / 100
        )
        st.caption(f"σ = 15% (volatilidad fija, típica de mercados accionarios)")

        n_simulaciones = st.select_slider(
            "Nº de simulaciones", [100, 250, 500, 1000, 2000], value=1000
        )
        meta_capital = st.number_input(
            "Meta de capital final (USD, opcional)",
            min_value=0, value=300000, step=10000, format="%d",
        )

        rerun = st.button("▶ SIMULAR / Re-aleatorizar", use_container_width=True)
        if "mc_seed" not in st.session_state or rerun:
            st.session_state["mc_seed"] = int(np.random.randint(0, 1_000_000))

    seed = st.session_state["mc_seed"]
    sigma = 0.15

    with st.spinner(f"Ejecutando {n_simulaciones} trayectorias..."):
        trayectorias, retornos_anuales = simular_monte_carlo(
            capital_inicial, aporte_mensual, anios,
            retorno_promedio, sigma=sigma, n_simulaciones=n_simulaciones, seed=seed,
        )

    # Trayectoria base determinística (retorno fijo)
    df_base = calcular_proyeccion(capital_inicial, aporte_mensual, anios, retorno_promedio, capitalizacion)
    saldo_base_meses = np.concatenate(([capital_inicial], df_base["saldo"].to_numpy()))
    saldo_base_final = float(saldo_base_meses[-1])

    # Sequence-of-returns: clasificar por promedio de los primeros 10 años
    # (si plazo < 10, usa los primeros plazo//2 años, mínimo 3)
    n_anios_inicio = max(3, min(10, anios // 2 if anios >= 6 else anios))
    promedio_inicio = retornos_anuales[:, :n_anios_inicio].mean(axis=1)
    q1 = np.percentile(promedio_inicio, 25)
    q3 = np.percentile(promedio_inicio, 75)

    idx_pesimistas = promedio_inicio <= q1   # cuartil inferior por inicio (malos primeros años)
    idx_optimistas = promedio_inicio >= q3   # cuartil superior por inicio (buenos primeros años)

    traj_pesimistas = trayectorias[idx_pesimistas]
    traj_optimistas = trayectorias[idx_optimistas]

    # Línea pesimista = mediana de las trayectorias del cuartil inferior
    linea_pesimista = np.median(traj_pesimistas, axis=0)
    linea_optimista = np.median(traj_optimistas, axis=0)

    cap_pesimista = float(linea_pesimista[-1])
    cap_optimista = float(linea_optimista[-1])

    # Probabilidades
    finales = trayectorias[:, -1]
    prob_meta = float((finales >= meta_capital).mean() * 100) if meta_capital > 0 else None
    prob_supera_base = float((finales >= saldo_base_final).mean() * 100)

    # Persistir resultados para el PDF
    st.session_state["mc_pesimista"] = cap_pesimista
    st.session_state["mc_base"] = saldo_base_final
    st.session_state["mc_optimista"] = cap_optimista
    st.session_state["mc_prob_meta"] = prob_meta
    st.session_state["mc_prob_base"] = prob_supera_base

    with col2:
        cA, cB, cC = st.columns([1, 1.2, 1])
        with cA:
            st.markdown(
                f'<div class="kpi-card" style="border-top:4px solid {BDI_GRIS_SUAVE};">'
                f'<div class="kpi-label">PESIMISTA</div>'
                f'<div style="font-size:10px; color:{BDI_GRIS_SUAVE}; font-style:italic; margin-bottom:3px;">malos primeros {n_anios_inicio} años</div>'
                f'<div class="kpi-value" style="color:{BDI_GRIS_SUAVE};">${cap_pesimista:,.0f}</div>'
                f'<div style="font-size:11px; color:{BDI_GRIS_SUAVE};">cuartil inferior por inicio</div></div>',
                unsafe_allow_html=True,
            )
        with cB:
            st.markdown(
                f'<div class="kpi-card kpi-card-primary" style="border-top:4px solid {BDI_LIMA};">'
                f'<div class="kpi-label">BASE · {retorno_promedio*100:.1f}% FIJO</div>'
                f'<div style="font-size:10px; color:{BDI_LIMA}; font-style:italic; margin-bottom:3px;">retorno constante</div>'
                f'<div class="kpi-value">${saldo_base_final:,.0f}</div>'
                f'<div style="font-size:11px; color:{BDI_LIMA};">trayectoria determinística</div></div>',
                unsafe_allow_html=True,
            )
        with cC:
            st.markdown(
                f'<div class="kpi-card" style="border-top:4px solid {BDI_CYAN};">'
                f'<div class="kpi-label">OPTIMISTA</div>'
                f'<div style="font-size:10px; color:{BDI_GRIS_SUAVE}; font-style:italic; margin-bottom:3px;">buenos primeros {n_anios_inicio} años</div>'
                f'<div class="kpi-value" style="color:{BDI_CYAN};">${cap_optimista:,.0f}</div>'
                f'<div style="font-size:11px; color:{BDI_GRIS_SUAVE};">cuartil superior por inicio</div></div>',
                unsafe_allow_html=True,
            )

        # Probabilidades
        cA2, cB2 = st.columns(2)
        if prob_meta is not None and meta_capital > 0:
            with cA2:
                st.markdown(
                    f"""
                    <div style="background:linear-gradient(90deg, {BDI_LIMA}, {BDI_CYAN}); padding:10px 14px; border-radius:8px; margin-top:10px; display:flex; align-items:center; gap:10px;">
                        <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:30px; color:{BDI_VERDE}; letter-spacing:1px;">{prob_meta:.0f}%</div>
                        <div style="font-size:12px; color:{BDI_VERDE};">superan tu meta de <strong>${meta_capital:,.0f}</strong></div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        with cB2:
            st.markdown(
                f"""
                <div style="background:{BDI_CREMA}; padding:10px 14px; border-radius:8px; margin-top:10px; display:flex; align-items:center; gap:10px; border-left:4px solid {BDI_VERDE};">
                    <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:30px; color:{BDI_VERDE}; letter-spacing:1px;">{prob_supera_base:.0f}%</div>
                    <div style="font-size:12px; color:{BDI_GRIS};">superan la base de <strong>${saldo_base_final:,.0f}</strong></div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # Gráfico
    st.markdown(f'<div class="section-title">TRAYECTORIAS SIMULADAS</div>', unsafe_allow_html=True)

    meses = np.arange(trayectorias.shape[1])
    p10_t = np.percentile(trayectorias, 10, axis=0)
    p90_t = np.percentile(trayectorias, 90, axis=0)
    anios_eje = meses / 12

    fig = go.Figure()
    # Banda de incertidumbre P10-P90 (lima translúcido)
    fig.add_trace(go.Scatter(
        x=np.concatenate([anios_eje, anios_eje[::-1]]),
        y=np.concatenate([p90_t, p10_t[::-1]]),
        fill="toself", fillcolor="rgba(181, 230, 29, 0.20)",
        line=dict(color="rgba(0,0,0,0)"),
        name="Banda P10–P90 (Monte Carlo)", hoverinfo="skip",
    ))
    # Línea pesimista (gris punteado)
    fig.add_trace(go.Scatter(
        x=anios_eje, y=linea_pesimista,
        line=dict(color=BDI_GRIS_SUAVE, width=2.5, dash="dash"),
        name=f"Pesimista (malos primeros {n_anios_inicio} años)",
        hovertemplate="Año %{x:.1f}<br>Pesimista: $%{y:,.0f}<extra></extra>",
    ))
    # Línea base determinística (verde sólido grueso)
    fig.add_trace(go.Scatter(
        x=anios_eje, y=saldo_base_meses,
        line=dict(color=BDI_VERDE, width=3),
        name=f"Base ({retorno_promedio*100:.1f}% fijo)",
        hovertemplate="Año %{x:.1f}<br>Base: $%{y:,.0f}<extra></extra>",
    ))
    # Línea optimista (cyan sólida)
    fig.add_trace(go.Scatter(
        x=anios_eje, y=linea_optimista,
        line=dict(color=BDI_CYAN, width=2.5),
        name=f"Optimista (buenos primeros {n_anios_inicio} años)",
        hovertemplate="Año %{x:.1f}<br>Optimista: $%{y:,.0f}<extra></extra>",
    ))
    if meta_capital > 0:
        fig.add_hline(
            y=meta_capital, line_dash="dot", line_color=BDI_GRIS_SUAVE,
            annotation_text=f"Meta ${meta_capital:,.0f}", annotation_position="top right",
        )
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=10, b=10, l=10, r=10),
        font=dict(family="Poppins", color=BDI_GRIS),
        xaxis=dict(title="Años", showgrid=False),
        yaxis=dict(title="USD", tickformat="$,.0f", gridcolor=BDI_CREMA),
        legend=dict(orientation="h", y=-0.2),
        height=440,
    )
    st.plotly_chart(fig, use_container_width=True)

    # Caja educativa al pie
    st.markdown(
        f"""
        <div style="background:{BDI_CREMA}; border-radius:10px; padding:14px; margin-top:12px;">
            <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:16px; color:{BDI_VERDE}; letter-spacing:1.5px; margin-bottom:6px;">
                📚 ¿POR QUÉ HAY DIFERENCIAS SI EL PROMEDIO ES EL MISMO?
            </div>
            <div style="font-size:12px; color:{BDI_GRIS}; line-height:1.6;">
                Cuando aportás <strong>todos los meses</strong>, importa CUÁNDO te tocan los malos años. Sortear 1.000 trayectorias aleatorias con el mismo promedio del {retorno_promedio*100:.1f}% deja en evidencia el efecto de la <strong>secuencia de retornos</strong>.
            </div>
            <div style="font-size:11px; color:{BDI_GRIS_SUAVE}; line-height:1.6; margin-top:8px; font-style:italic;">
                Dato contraintuitivo: en fase de acumulación, el escenario pesimista (malos primeros años) a veces termina mejor que el optimista, porque comprás barato al inicio y aprovechás la recuperación. Es el mismo efecto que hace funcionar al dollar-cost averaging.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# TAB 4 - APRENDÉ (educativo)
# ============================================================
def tab_aprende() -> None:
    st.markdown(
        f"""
        <div style="background:linear-gradient(135deg, {BDI_VERDE}, {BDI_CYAN}); border-radius:12px; padding:24px; color:white; margin:14px 0;">
            <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:32px; letter-spacing:2px; line-height:1;">ANTES DE PROYECTAR, ENTENDÉ</div>
            <div style="font-size:13px; color:{BDI_CREMA}; margin-top:8px;">Toda calculadora hace supuestos. Acá te contamos cuáles, cómo afectan los números y dónde están los límites del modelo.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    conceptos = [
        ("🪙", "INTERÉS COMPUESTO",
         "Cada período el rendimiento se reinvierte y empieza a generar su propio rendimiento. Es la base del crecimiento exponencial.",
         "FV = P(1+r/n)^(n·t)", BDI_VERDE),
        ("📅", "CAPITALIZACIÓN",
         "Cuántas veces por año se capitaliza el interés. A mayor frecuencia, ligeramente más rendimiento.",
         "Anual=1 · Mensual=12 · Diaria=365", BDI_CYAN),
        ("📊", "RETORNO ESPERADO",
         "El default de 8% es el promedio histórico de carteras balanceadas. La realidad es volátil año a año.",
         "S&P 500 hist. ≈ 10% nominal", BDI_LIMA),
        ("🎯", "REGLA DEL 4%",
         "Bengen (1994): podés retirar el 4% del capital al año sin agotarlo en 30 años de retiro.",
         "Capital = ingreso anual / 0.04", BDI_VERDE),
        ("🎲", "MONTE CARLO",
         "Simula miles de futuros con rendimientos aleatorios. Da rangos de probabilidad, no certezas absolutas.",
         "Útil para cuantificar riesgo", BDI_CYAN),
        ("⚠", "SUPUESTOS DEL MODELO",
         "Sin inflación · sin impuestos · sin comisiones · aportes constantes · retorno fijo (salvo Monte Carlo).",
         "Realidad ≠ proyección", BDI_LIMA),
    ]
    cols = st.columns(3)
    for i, (icon, titulo, desc, formula, color) in enumerate(conceptos):
        with cols[i % 3]:
            st.markdown(
                f"""
                <div style="background:{BDI_CREMA}; border-radius:10px; padding:16px; border-top:4px solid {color}; margin-bottom:12px; min-height:200px;">
                    <div style="font-size:28px; margin-bottom:4px;">{icon}</div>
                    <div style="font-family:'Bebas Neue', Impact, sans-serif; font-size:18px; color:{BDI_VERDE}; letter-spacing:1px;">{titulo}</div>
                    <div style="font-size:13px; color:{BDI_GRIS}; line-height:1.55; margin-top:8px;">{desc}</div>
                    <div style="background:white; padding:6px 10px; border-radius:4px; font-family:monospace; font-size:12px; color:{BDI_VERDE}; margin-top:10px;">{formula}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # Mini quiz
    st.markdown(f'<div class="section-title">🧠 MINI QUIZ · ¿LO ENTENDISTE?</div>', unsafe_allow_html=True)
    pregunta = st.radio(
        "Si invertís $10.000 al 8% durante 30 años con capitalización mensual, ¿cuánto tenés al final?",
        ["a) $54.000", "b) $80.000", "c) $109.357", "d) $240.000"],
        index=None,
    )
    if pregunta == "c) $109.357":
        st.success("✅ ¡Correcto! El interés compuesto hace casi todo el trabajo. Solo el último año generás ~$8.000 sin aportar un dólar más.")
    elif pregunta is not None:
        st.warning("Probá de nuevo. Pista: a 8% durante 30 años el dinero se multiplica por más de 10x.")

    # Disclaimer
    st.markdown(
        f"""
        <div class="disclaimer">
            <strong>⚠ Disclaimer BDI:</strong> Esta calculadora es una herramienta educativa. Las proyecciones no garantizan rendimientos futuros. Consultá siempre con tu asesor BDI antes de tomar decisiones de inversión.
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# MAIN
# ============================================================
def main() -> None:
    inject_css()
    render_header()

    capital_inicial, aporte_mensual, anios, retorno, capitalizacion = render_sidebar()

    tab1, tab2, tab3, tab4 = st.tabs([
        "1 · CLÁSICA",
        "2 · METAS",
        "3 · ESCENARIOS",
        "📚 APRENDÉ",
    ])

    with tab1:
        tab_clasica(capital_inicial, aporte_mensual, anios, retorno, capitalizacion)

    with tab2:
        tab_metas(capital_inicial, aporte_mensual, anios, retorno, capitalizacion)

    with tab3:
        tab_montecarlo(capital_inicial, aporte_mensual, anios, retorno, capitalizacion)

    with tab4:
        tab_aprende()


if __name__ == "__main__":
    main()
